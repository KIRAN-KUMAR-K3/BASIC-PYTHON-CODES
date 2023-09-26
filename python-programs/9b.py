"""
py -m pip install requests
py -m pip install openpyxl"""
import openpyxl
path = "E:/gfg.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj['A1': 'B6']
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "ANKIT"
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "RAI"
c3 = sheet['A2']
c3.value = "RAHUL"
c4 = sheet['B2']
c4.value = "RAI"
wb.save("E:\\demo.xlsx")